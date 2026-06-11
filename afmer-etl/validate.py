#!/usr/bin/env python3
"""Validate the Rust ETL output against the reference AFMER-*.xlsx workbooks.

For every report year present in both the SQLite database and an AFMER-<year>.xlsx,
compares row coverage, the 22 numeric production/export counts (which must match
exactly), the key/type/state fields, and the license-name / premise-street split
(for which the XLSX is ground truth — divergences here come from the FFL-based
name split and are reported, not failed).

Usage: .venv/bin/python afmer-etl/validate.py [afmer.db] [SOURCE/NU dir]
"""
import sys, os, re, sqlite3, glob
import openpyxl

DB = sys.argv[1] if len(sys.argv) > 1 else "afmer.db"
XLSX_DIR = sys.argv[2] if len(sys.argv) > 2 else "SOURCE/NU"

COUNT_COLS = [
    "PSTL_22","PSTL_25","PSTL_32","PSTL_380","PSTL_9MM","PSTL_50","PSTL_TOTL",
    "RVLR_22","RVLR_32","RVLR_38","RVLR_357","RVLR_44","RVLR_50","RVLR_TOTL",
    "RIFLE_MFG","SHOTGUN_MFG","MIS_FAM","PISTOL_EXP","REVOLVER_EXP","RIFLE_EXP",
    "SHOTGUN_EXP","MISC_FA_EXP",
]


def load_xlsx(path):
    """Return {rds_key: dict(header->value)} keyed by header names from row 1."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip().upper() if h is not None else "" for h in next(rows)]
    out = {}
    for r in rows:
        rec = {h: ("" if v is None else str(v).strip()) for h, v in zip(headers, r)}
        key = rec.get("APP_RDS_KEY", "")
        if key:
            out[key] = rec
    wb.close()
    return out


def count_xlsx_rows(path):
    """Total data rows (including duplicate keys), excluding header."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    n = 0
    it = ws.iter_rows(values_only=True)
    next(it, None)  # header
    for r in it:
        if r and r[0] not in (None, ""):
            n += 1
    wb.close()
    return n


def main():
    if not os.path.exists(DB):
        sys.exit(f"no database at {DB}")
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    years = [y for (y,) in conn.execute("SELECT DISTINCT source_year FROM afmer ORDER BY source_year DESC")]
    if not years:
        sys.exit("no afmer rows in database")

    overall_ok = True
    for year in years:
        xlsx = os.path.join(XLSX_DIR, f"AFMER-{year}.xlsx")
        # XLSX files also live in the repo root for some years.
        if not os.path.exists(xlsx):
            alt = f"AFMER-{year}.xlsx"
            xlsx = alt if os.path.exists(alt) else xlsx
        print(f"\n=== {year} ===")
        if not os.path.exists(xlsx):
            print(f"  no reference XLSX ({xlsx}); skipping")
            continue

        gt = load_xlsx(xlsx)
        db_total = conn.execute(
            "SELECT COUNT(*) FROM afmer WHERE source_year=?", (year,)).fetchone()[0]
        xl_total = count_xlsx_rows(xlsx)
        print(f"  raw rows (incl. duplicate keys): db={db_total} xlsx={xl_total}")
        db_rows = {r["rds_key"]: r for r in
                   conn.execute("SELECT * FROM afmer WHERE source_year=?", (year,))}

        only_db = set(db_rows) - set(gt)
        only_xl = set(gt) - set(db_rows)
        common = set(db_rows) & set(gt)
        print(f"  rows: db={len(db_rows)} xlsx={len(gt)} common={len(common)} "
              f"only_db={len(only_db)} only_xlsx={len(only_xl)}")

        counts_ok = type_ok = state_ok = name_ok = street_ok = 0
        count_mismatch_samples = []
        for k in common:
            d, g = db_rows[k], gt[k]
            if d["app_lic_type"] == g.get("APP_LIC_TYPE", ""):
                type_ok += 1
            if (d["app_premise_state"] or "") == g.get("APP_PREMISE_STATE", ""):
                state_ok += 1
            if (d["app_license_name"] or "").upper() == g.get("APP_LICENSE_NAME", "").upper():
                name_ok += 1
            if (d["app_premise_city"] or "").upper() == g.get("APP_PREMISE_CITY", "").upper():
                street_ok += 1
            cm = True
            for col in COUNT_COLS:
                dv = d[col.lower()]
                gv = g.get(col, "0")
                try:
                    if int(dv) != int(float(gv or 0)):
                        cm = False
                        break
                except ValueError:
                    cm = False
                    break
            if cm:
                counts_ok += 1
            elif len(count_mismatch_samples) < 5:
                count_mismatch_samples.append(k)

        n = len(common) or 1
        def pct(x): return f"{x}/{len(common)} ({100*x/n:.2f}%)"
        print(f"  counts exact : {pct(counts_ok)}")
        print(f"  lic_type     : {pct(type_ok)}")
        print(f"  state        : {pct(state_ok)}")
        print(f"  license_name : {pct(name_ok)}   (XLSX is ground truth; rest from FFL split)")
        print(f"  premise_street(city col): {pct(street_ok)}")
        if count_mismatch_samples:
            print(f"  count-mismatch sample keys: {count_mismatch_samples}")
            k = count_mismatch_samples[0]
            d, g = db_rows[k], gt[k]
            diffs = [(c, d[c.lower()], g.get(c)) for c in COUNT_COLS
                     if str(d[c.lower()]) != str(int(float(g.get(c) or 0)))]
            print(f"    {k} diffs (col, db, xlsx): {diffs}")

        # Pass criteria: every common row's counts match exactly, and coverage
        # of the XLSX is complete.
        if counts_ok != len(common) or only_xl:
            overall_ok = False

    print("\nRESULT:", "PASS" if overall_ok else "REVIEW (see above)")
    conn.close()


if __name__ == "__main__":
    main()
